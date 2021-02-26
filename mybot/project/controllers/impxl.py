import os
import pprint
import json
from openpyxl import load_workbook

from mybot.config import RESOURCES_PATH

file_path = [RESOURCES_PATH, 'settings', 'logo.xlsx']
logo = os.path.join(*file_path)

# logo = 'logo.xlsx'
newDict = {}


def load_city(book):

    row = 2
    sheet = book['city']
    id_city = {}

    for i in range(row, sheet.max_row + 1):
        idx_ = sheet.cell(column=1, row=i).value
        city_ = sheet.cell(column=2, row=i).value
        access_ = sheet.cell(column=3, row=i).value

        if idx_ in id_city:
            t = id_city[idx_]
            if access_:
                t[2].append(access_)
            id_city[idx_] = t
        else:
            if access_:
                id_city[idx_] = [idx_, city_, [access_]]
            else:
                id_city[idx_] = [idx_, city_, []]

    out_city = [id_city[b] for b in id_city]

    return out_city


def load_delivery(book):

    row = 2
    sheet = book['delivery']
    id_dlv = {}

    for i in range(row, sheet.max_row + 1):
        idx_ = sheet.cell(column=1, row=i).value
        city_ = sheet.cell(column=2, row=i).value
        access_ = sheet.cell(column=3, row=i).value

        if idx_ in id_dlv:
            t = id_dlv[idx_]
            if access_:
                t[2].append(access_)
            id_dlv[idx_] = t
        else:
            if access_:
                id_dlv[idx_] = [idx_, city_, [access_]]
            else:
                id_dlv[idx_] = [idx_, city_, []]

    out_dlv = [id_dlv[b] for b in id_dlv]

    return out_dlv


def load_weight(book):

    row = 2
    sheet = book['weight']
    id_weight = {}

    for i in range(row, sheet.max_row + 1):
        idx_ = sheet.cell(column=1, row=i).value
        city_ = sheet.cell(column=2, row=i).value
        access_ = sheet.cell(column=3, row=i).value

        if idx_ in id_weight:
            t = id_weight[idx_]
            if access_:
                t[2].append(access_)
            id_weight[idx_] = t
        else:
            if access_:
                id_weight[idx_] = [idx_, city_, [access_]]
            else:
                id_weight[idx_] = [idx_, city_, []]

    out_weight = [id_weight[b] for b in id_weight]

    return out_weight


def load_address(book):

    row = 2
    sheet = book['adr']
    combine = {}

    for i in range(row, sheet.max_row + 1):
        id_city = sheet.cell(column=1, row=i).value
        id_adr = sheet.cell(column=2, row=i).value

        city_adr = str(id_city) + str(id_adr)

        adr_name = sheet.cell(column=3, row=i).value
        access_ = sheet.cell(column=4, row=i).value

        if city_adr in combine:
            t = combine[city_adr]
            if access_:
                t[3].append(access_)
            combine[city_adr] = t
        else:
            if access_:
                combine[city_adr] = [id_city, id_adr, adr_name, [access_]]
            else:
                combine[city_adr] = [id_city, id_adr, adr_name, []]

    out_adr = [combine[b] for b in combine]

    return out_adr


def run():
    wb = load_workbook(logo)
    newDict['city'] = load_city(wb)
    newDict['delivery'] = load_delivery(wb)
    newDict['wt'] = load_weight(wb)
    newDict['adr'] = load_address(wb)
    wb.close()


if __name__ == '__main__':

    run()

    file_path = [RESOURCES_PATH, 'settings', 'data.txt']
    djs = os.path.join(*file_path)
    with open(djs, 'w') as outfile:
        json.dump(newDict, outfile)

    pprint.pprint(newDict)
