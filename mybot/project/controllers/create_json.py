
# Импорт массивов из файла

import os
import pprint
import json
from openpyxl import load_workbook

from mybot.config import RESOURCES_PATH

newDict = {}


def create_cities(ws, adr):
    """ Create Dict with cities name """
    tp = []

    for row in ws.iter_rows(min_row=adr['start'][0],
                            min_col=adr['start'][1],
                            max_row=adr['finish'][0],
                            max_col=adr['finish'][1]):
        type_list = []
        for b in row:
            type_list.append(b.value)
        type_list.append([])
        tp.append(type_list)
    return tp


def load_sheet(wb):

    address_data = [
        {'start': [2, 7], 'finish': [17, 8]},
        {'start': [2, 1], 'finish': [43, 3]},
        {'start': [22, 7], 'finish': [24, 8]},
        {'start': [29, 7], 'finish': [30, 8]}
    ]

    for b in wb.sheetnames:
        ws = wb[b]
        newDict['city'] = create_cities(ws, address_data[0])
        newDict['adr'] = create_cities(ws, address_data[1])
        newDict['delivery'] = create_cities(ws, address_data[2])
        newDict['wt'] = create_cities(ws, address_data[3])


def open_book(pt):
    """ Открываем все книги по очереди
    """
    for filename in pt:
        tmpfile, fileExtension = os.path.splitext(filename)
        if '.xlsx' == fileExtension:
            wb = load_workbook(filename)
            load_sheet(wb)
            wb.close()


def run(rootDir=r'myjson'):
    for dirName, subdirList, fileList in os.walk(rootDir,
                                                 topdown=False):
        path_list = dirName.split(os.sep)
        pt = [os.path.join(".", dirName, b) for b in fileList]
        open_book(pt)


if __name__ == '__main__':
    run()

    # with open('data.txt', 'w') as outfile:
    #     json.dump(newDict, outfile)

    file_path = [RESOURCES_PATH, 'settings', 'data.txt']
    djs = os.path.join(*file_path)

    with open(djs) as json_file:
        newDict = json.load(json_file)

    pprint.pprint(newDict)