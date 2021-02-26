
import os
import pprint
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from mybot.config import RESOURCES_PATH

newDict = {}

file_path = [RESOURCES_PATH, 'settings', 'tree.txt']
djs = os.path.join(*file_path)


def import_json():
    with open(djs) as json_file:
        newDict = json.load(json_file)

    pprint.pprint(newDict)
    return newDict


wb = Workbook()


def exp_weight(sheet, wt):

    _ = sheet.cell(column=1, row=1, value=f"id".format(get_column_letter(1)))
    _ = sheet.cell(column=2, row=1, value=f"weight".format(get_column_letter(1)))
    _ = sheet.cell(column=3, row=1, value=f"access".format(get_column_letter(1)))

    row = 2

    for b in wt:
        if b[2]:
            for access in b[2]:
                print(b[0], b[1], access)
                _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
                _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
                _ = sheet.cell(column=3, row=row, value=f"{access}".format(get_column_letter(3)))
                row += 1
        else:
            print(b[0], b[1], None)
            _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
            _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
            # _ = sheet.cell(column=3, row=row, value=f"None".format(get_column_letter(3)))

            row += 1


def exp_delivery(sheet, delivery):

    _ = sheet.cell(column=1, row=1, value=f"id".format(get_column_letter(1)))
    _ = sheet.cell(column=2, row=1, value=f"delivery".format(get_column_letter(1)))
    _ = sheet.cell(column=3, row=1, value=f"access".format(get_column_letter(1)))

    row = 2

    for b in delivery:
        if b[2]:
            for access in b[2]:
                print(b[0], b[1], access)
                _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
                _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
                _ = sheet.cell(column=3, row=row, value=f"{access}".format(get_column_letter(3)))
                row += 1
        else:
            print(b[0], b[1], None)
            _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
            _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
            # _ = sheet.cell(column=3, row=row, value=f"None".format(get_column_letter(3)))

            row += 1


def exp_adr(sheet, adr):

    _ = sheet.cell(column=1, row=1, value=f"id_city".format(get_column_letter(1)))
    _ = sheet.cell(column=2, row=1, value=f"id_adr".format(get_column_letter(1)))
    _ = sheet.cell(column=3, row=1, value=f"address".format(get_column_letter(1)))
    _ = sheet.cell(column=4, row=1, value=f"access".format(get_column_letter(1)))

    row = 2

    for b in adr:
        if b[3]:
            for access in b[3]:
                print(b[0], b[1], b[2], access)
                _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
                _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
                _ = sheet.cell(column=3, row=row, value=f"{b[2]}".format(get_column_letter(3)))
                _ = sheet.cell(column=4, row=row, value=f"{access}".format(get_column_letter(4)))
                row += 1
        else:
            print(b[0], b[1], b[2], None)
            _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
            _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
            _ = sheet.cell(column=3, row=row, value=f"{b[2]}".format(get_column_letter(3)))
            # _ = sheet.cell(column=4, row=row, value=f"None".format(get_column_letter(4)))

            row += 1


def exp_city(sheet, city):

    _ = sheet.cell(column=1, row=1, value=f"id".format(get_column_letter(1)))
    _ = sheet.cell(column=2, row=1, value=f"city".format(get_column_letter(1)))
    _ = sheet.cell(column=3, row=1, value=f"access".format(get_column_letter(1)))

    row = 2

    for b in city:
        if b[2]:
            for access in b[2]:
                print(b[0], b[1], access)
                _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
                _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
                _ = sheet.cell(column=3, row=row, value=f"{access}".format(get_column_letter(3)))
                row += 1
        else:
            print(b[0], b[1], None)
            _ = sheet.cell(column=1, row=row, value=f"{b[0]}".format(get_column_letter(1)))
            _ = sheet.cell(column=2, row=row, value=f"{b[1]}".format(get_column_letter(2)))
            # _ = sheet.cell(column=3, row=row, value=f"None".format(get_column_letter(3)))

            row += 1


def iter_dict(data):

    ws1 = wb.create_sheet(title='city')
    exp_city(ws1, data['city'])

    ws2 = wb.create_sheet(title='adr')
    exp_adr(ws2, data['adr'])

    ws3 = wb.create_sheet(title='delivery')
    exp_city(ws3, data['delivery'])

    ws4 = wb.create_sheet(title='weight')
    exp_city(ws4, data['wt'])

    _path = [RESOURCES_PATH, 'settings', 'logo.xlsx']
    excel_path = os.path.join(*_path)
    wb.save(excel_path)


if __name__ == "__main__":
    iter_dict(import_json())
